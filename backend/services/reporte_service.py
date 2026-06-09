import io
import datetime as dt

from dao import recibo_dao, gasto_dao
from services.recibo_service import get_morosos_prioridad
from utils.money import money_float, to_decimal
from utils.logger import get_logger

log = get_logger(__name__)


def get_resumen_financiero(mes: str) -> dict:
    rows = recibo_dao.get_recibos_admin(mes_filter=mes)

    total_emitido = 0.0
    total_cobrado = 0.0
    concepto_admin = 0.0
    concepto_agua = 0.0
    concepto_luz = 0.0
    concepto_mant = 0.0

    for row in rows:
        total = money_float(
            to_decimal(row["monto_administracion"])
            + to_decimal(row["monto_agua"])
            + to_decimal(row["monto_luz"])
            + to_decimal(row["monto_mantenimiento"])
        )
        pagado = money_float(row.get("monto_pagado") or 0)
        total_emitido += total
        total_cobrado += pagado
        concepto_admin += money_float(row["monto_administracion"])
        concepto_agua += money_float(row["monto_agua"])
        concepto_luz += money_float(row["monto_luz"])
        concepto_mant += money_float(row["monto_mantenimiento"])

    saldo_pendiente = round(total_emitido - total_cobrado, 2)
    porcentaje_cobranza = round((total_cobrado / total_emitido * 100), 2) if total_emitido > 0 else 0.0

    return {
        "mes": mes,
        "total_emitido": round(total_emitido, 2),
        "total_cobrado": round(total_cobrado, 2),
        "saldo_pendiente": round(saldo_pendiente, 2),
        "porcentaje_cobranza": porcentaje_cobranza,
        "desglose": {
            "administracion": round(concepto_admin, 2),
            "agua": round(concepto_agua, 2),
            "luz": round(concepto_luz, 2),
            "mantenimiento": round(concepto_mant, 2),
        },
    }


def exportar_morosidad_excel(mes: str = "", limit: int = 100) -> bytes:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise RuntimeError("openpyxl no está instalado")

    data = get_morosos_prioridad(mes, limit)
    items = data["items"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte Morosidad"

    # Encabezado del reporte
    ws.merge_cells("A1:H1")
    ws["A1"] = f"REPORTE DE MOROSIDAD - {mes or 'Todos los meses'}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A2"] = f"Generado: {dt.date.today().isoformat()}"
    ws["A2"].font = Font(italic=True, size=10)

    # Cabeceras
    headers = ["Ranking", "Propietario", "DNI/Depto", "Torre", "Saldo (S/)", "Días Pendiente", "Fecha Emisión", "Recibo ID"]
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Datos
    alt_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    for i, item in enumerate(items, start=1):
        row_idx = i + 4
        fill = alt_fill if i % 2 == 0 else None
        propietario = item.get("propietario") or {}
        nombre = f"{propietario.get('nombre', '')} {propietario.get('apellido', '')}".strip()
        values = [
            i,
            nombre,
            item.get("nro_departamento", ""),
            item.get("torre", ""),
            item.get("saldo", 0),
            item.get("dias_pendiente", 0),
            item.get("fecha_emision", ""),
            item.get("recibo_id", ""),
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            if fill:
                cell.fill = fill

    # Resumen al final
    total_filas = len(items)
    sum_row = total_filas + 6
    ws.cell(row=sum_row, column=1, value="TOTAL MOROSOS").font = Font(bold=True)
    ws.cell(row=sum_row, column=2, value=total_filas)
    ws.cell(row=sum_row + 1, column=1, value="DEUDA TOTAL (S/)").font = Font(bold=True)
    ws.cell(row=sum_row + 1, column=2, value=round(sum(i.get("saldo", 0) for i in items), 2))

    # Ajustar ancho de columnas (saltar MergedCells que no tienen column_letter)
    for col in ws.columns:
        try:
            letter = col[0].column_letter
        except AttributeError:
            continue
        max_len = max((len(str(cell.value or "")) for cell in col if hasattr(cell, 'value')), default=10)
        ws.column_dimensions[letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    log.info("Excel morosidad generado", extra={"filas": total_filas})
    return buf.read()
